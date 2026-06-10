import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
import textwrap
import re
import requests
import io
from google.oauth2 import service_account
import google.auth.transport.requests

# ==================================================
# PAGE CONFIG
# ==================================================
st.set_page_config(
    page_title="PEA Executive Dashboard",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================================================
# SESSION STATE INITIALIZATION
# ==================================================
if 'app_state' not in st.session_state:
    st.session_state.app_state = 'connect' 
if 'summary_df' not in st.session_state:
    st.session_state.summary_df = None
if 'all_df' not in st.session_state:
    st.session_state.all_df = None
if 'gsheet_url' not in st.session_state:
    st.session_state.gsheet_url = ""

# ==================================================
# CSS GLOBAL
# ==================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sarabun:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"]  {
    font-family: 'Sarabun', sans-serif;
}

[data-testid="stAppViewContainer"] {
    background-color: #F8FAFC; 
    color: #1E293B; 
}
[data-testid="stHeader"] {
    background-color: transparent;
}
.block-container { padding-top: 2rem; padding-bottom: 2rem; }

.pea-header {
    background: linear-gradient(135deg, #0A192F, #4C1D95, #0284C7); 
    color: white;
    padding: 32px;
    border-radius: 20px;
    margin-bottom: 32px;
    box-shadow: 0 10px 25px -5px rgba(76, 29, 149, 0.4); 
    border: 1px solid rgba(255,255,255,0.1);
}
.pea-title { font-size: 36px; font-weight: 700; margin-bottom: 8px; color: #FFFFFF; letter-spacing: -0.5px;}
.pea-subtitle { font-size: 16px; font-weight: 400; color: #E0E7FF; letter-spacing: 0.5px;}

.glass-card {
    background: rgba(255, 255, 255, 0.9); 
    backdrop-filter: blur(12px);
    -webkit-backdrop-filter: blur(12px);
    border-radius: 20px;
    padding: 24px;
    box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.05), 0 4px 6px -2px rgba(0, 0, 0, 0.025); 
    border: 1px solid rgba(255, 255, 255, 0.6);
    transition: transform 0.3s ease, box-shadow 0.3s ease;
    margin-bottom: 24px;
}
.glass-card:hover {
    transform: translateY(-4px);
    box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04);
}

.kpi-title { font-size: 15px; color: #64748B; margin-bottom: 8px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;}
.kpi-value { font-size: 36px; font-weight: 700; line-height: 1.2; }

.category-card { min-height: 220px; }
.category-number { color: #3730A3; font-size: 24px; font-weight: 800; opacity: 0.2; float: right;}
.category-title { font-size: 18px; font-weight: 700; margin: 0 0 12px 0; color: #0F172A;}
.category-value { font-size: 15px; margin-bottom: 8px; color: #475569; font-weight: 500;}
.category-note { color: #0EA5E9; margin-top: 16px; font-size: 13px; border-top: 1px solid #F1F5F9; padding-top: 12px; font-weight: 500;}

.table-header-done { color: #0BFF8D; font-size: 18px; font-weight: 700; margin-top: 25px; border-bottom: 2px solid #0BFF8D; padding-bottom: 8px;}
.table-header-prog { color: #D58A19; font-size: 18px; font-weight: 700; margin-top: 25px; border-bottom: 2px solid #D58A19; padding-bottom: 8px;}
.table-header-not  { color: #F43F5E; font-size: 18px; font-weight: 700; margin-top: 25px; border-bottom: 2px solid #F43F5E; padding-bottom: 8px;}

/* Custom CSS HTML Table */
.styled-table {
    width: 100%;
    border-collapse: collapse;
    margin: 15px 0 20px 0;
    font-size: 14px;
    background-color: white;
    border: 1px solid #E2E8F0;
}
.styled-table thead tr {
    background-color: #F8FAFC;
    color: #475569;
}
.styled-table th {
    padding: 12px 15px;
    border: 1px solid #E2E8F0;
    text-align: center !important; 
    vertical-align: middle;
    white-space: nowrap; 
}
.styled-table td {
    padding: 12px 15px;
    border: 1px solid #E2E8F0;
    vertical-align: top;
    color: #334155;
    white-space: nowrap; 
}
.styled-table td:nth-child(2),
.styled-table td:nth-child(8) {
    white-space: normal !important; 
    word-wrap: break-word;
    min-width: 250px; 
}
.styled-table tbody tr:hover {
    background-color: #F1F5F9;
}

::-webkit-scrollbar { width: 8px; height: 8px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #CBD5E1; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #94A3B8; }
</style>
""", unsafe_allow_html=True)

# ==================================================
# API & DATA PROCESSING FUNCTIONS
# ==================================================
THAI_MONTHS = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]

def extract_year_from_text(text):
    if not text: return ""
    text = str(text).strip() 
    
    match = re.search(r'SCOD\s*(\d{2})', text, re.IGNORECASE)
    if match: return match.group(1)
    
    match = re.search(r'\((\d{2})\)', text)
    if match: return match.group(1)
    
    match = re.search(r'ปี\s*(25\d{2}|\d{2})', text)
    if match:
        year_str = match.group(1)
        return year_str[-2:] if len(year_str) == 4 else year_str
        
    match = re.search(r'(25\d{2})', text)
    if match: return match.group(1)[-2:]
    
    match = re.search(r'([6-9]\d)$', text)
    if match: return match.group(1)
    
    return ""

def fetch_private_gsheet_excel(url):
    try:
        match = re.search(r'\/d\/(.*?)\/', url)
        if not match:
            st.error("ลิงก์ไม่ถูกต้อง กรุณาตรวจสอบลิงก์ Google Sheets อีกครั้ง")
            return None
        sheet_id = match.group(1)

        creds_info = st.secrets["connections"]["gsheets"]
        credentials = service_account.Credentials.from_service_account_info(
            creds_info,
            scopes=['https://www.googleapis.com/auth/drive.readonly']
        )

        request = google.auth.transport.requests.Request()
        credentials.refresh(request)

        export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        headers = {'Authorization': f'Bearer {credentials.token}'}
        
        response = requests.get(export_url, headers=headers)
        
        if response.status_code == 200:
            return io.BytesIO(response.content)
        else:
            st.error(f"ไม่สามารถเข้าถึงไฟล์ได้ (Status {response.status_code}): โปรดตรวจสอบว่าคุณได้แชร์ไฟล์ให้กับอีเมลใน Service Account เป็น Viewer แล้วหรือยัง")
            return None

    except KeyError:
        st.error("ไม่พบข้อมูล API Key ในไฟล์ .streamlit/secrets.toml โปรดตั้งค่าให้เรียบร้อย")
        return None
    except Exception as e:
        st.error(f"เกิดข้อผิดพลาดในการโหลดข้อมูล: {e}")
        return None

@st.cache_data(show_spinner=False)
def get_all_excel_data(file_bytes):
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    all_data = []
    
    keywords_mapping = {
        "โครงการ": ["โครงการ", "ประเภทงาน", "หมวดงาน"],
        "ชื่องาน": ["ชื่องาน", "ชื่อโครงการย่อย", "รายละเอียดงาน", "รายการ"],
        "พื้นที่": ["พื้นที่", "กฟข.", "เขต"], 
        "จังหวัด": ["จังหวัด"],
        "ผู้รับผิดชอบ": ["ผู้รับผิดชอบ", "ผู้ดำเนินการ"],
        "วงเงิน": ["วงเงินตามอนุมัติ", "วงเงินงบประมาณ", "งบประมาณ", "วงเงิน"],
        "หมายเหตุ": ["อนุมัติ/หมายเหตุ/ปัญหา", "หมายเหตุ", "ปัญหา"],
        "สถานะ": ["สถานะงาน", "ตรวจสอบสถานะงาน", "ตรวจสอบงานแล้วเสร็จ", "สถานะ"]
    }

    for sheet_name in wb.sheetnames:
        if sheet_name == "สรุป": continue
        
        ws = wb[sheet_name]
        col_map = {}
        month_cols = {}
        header_row = 1
        
        for r in range(1, 15):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(r, c).value
                if isinstance(val, str):
                    val_str = val.strip()
                    val_lower = val_str.lower()
                    
                    for std_key, aliases in keywords_mapping.items():
                        if std_key not in col_map:
                            for alias in aliases:
                                if alias.lower() in val_lower:
                                    col_map[std_key] = c
                                    header_row = max(header_row, r)
                                    break
                    
                    if val_str in THAI_MONTHS:
                        month_cols[c] = val_str

        if "โครงการ" not in col_map: continue
        
        proj_c = col_map["โครงการ"]
        task_c = col_map.get("ชื่องาน", proj_c + 2) 

        current_project = ""

        for r in range(header_row + 1, ws.max_row + 1):
            has_data = False
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip() != "":
                    has_data = True
                    break
            if not has_data: continue

            project_val = ws.cell(r, proj_c).value
            if project_val is not None and str(project_val).strip() != "":
                current_project = str(project_val).strip()
                
            task_val = ws.cell(r, task_c).value
            task_str = str(task_val).strip() if task_val is not None else ""
            
            if not task_str or task_str in THAI_MONTHS or "ไตรมาส" in task_str:
                continue
            
            task_clean = task_str.replace(" ", "")
            proj_clean = current_project.replace(" ", "")
            if task_clean in ["รวม", "รวมทั้งสิ้น", "ยอดรวม", "รวมทั้งหมด"] or proj_clean in ["รวม", "รวมทั้งสิ้น", "ยอดรวม"]:
                continue
            
            year_suffix = extract_year_from_text(current_project)
            if not year_suffix:
                year_suffix = extract_year_from_text(sheet_name)
            
            deadline_str = "-"
            sort_index = 9999 
            
            for c in sorted(month_cols.keys(), reverse=True):
                cell = ws.cell(r, c)
                if cell.fill and cell.fill.patternType == 'solid':
                    color = cell.fill.start_color.rgb
                    if color and str(color) not in ['00000000', 'FFFFFFFF', '000000', 'None']:
                        month_str = month_cols[c]
                        deadline_str = f"{month_str} {year_suffix}".strip() if year_suffix else month_str
                        sort_index = c 
                        break 
            
            row = {
                "แหล่งที่มา (ชีต)": sheet_name,
                "โครงการหลัก": current_project,
                "ชื่องาน": task_str,
                "พื้นที่": str(ws.cell(r, col_map["พื้นที่"]).value or "-").strip() if "พื้นที่" in col_map else "-",
                "จังหวัด": str(ws.cell(r, col_map["จังหวัด"]).value or "-").strip() if "จังหวัด" in col_map else "-",
                "ผู้รับผิดชอบ": str(ws.cell(r, col_map["ผู้รับผิดชอบ"]).value or "-").strip() if "ผู้รับผิดชอบ" in col_map else "-",
                "วงเงิน": ws.cell(r, col_map["วงเงิน"]).value if "วงเงิน" in col_map else 0,
                "หมายเหตุ/ปัญหา": str(ws.cell(r, col_map["หมายเหตุ"]).value or "-").strip() if "หมายเหตุ" in col_map else "-"
            }
            
            raw_status = ws.cell(r, col_map["สถานะ"]).value if "สถานะ" in col_map else None
            row["สถานะดิบ"] = str(raw_status).strip() if raw_status is not None else ""
            row["Deadline"] = deadline_str
            row["Sort_Index"] = sort_index 
            
            all_data.append(row)

    df_combined = pd.DataFrame(all_data)
    
    if not df_combined.empty:
        # ---- ทำความสะอาดข้อมูล: ตัดคำว่า (เพิ่มเติม) ออกจากโครงการหลัก ----
        df_combined["โครงการหลัก"] = df_combined["โครงการหลัก"].astype(str).str.replace(r'\s*\(เพิ่มเติม\)', '', regex=True)
        
        def map_3_status(val):
            val = str(val).strip().lower()
            if val in ["nan", "none", "", "nat", "-", "0", "0.0"]: return "❌ ยังไม่ได้ดำเนินการ"
            if "ยังไม่ได้" in val or "ยังไม่เริ่ม" in val: return "❌ ยังไม่ได้ดำเนินการ"
            if any(x in val for x in ["1", "1.0", "สำเร็จ", "แล้วเสร็จ", "เสร็จ"]): return "✅ แล้วเสร็จ"
            if any(x in val for x in ["กำลัง", "ระหว่าง", "ดำเนินการ"]): return "⏳ อยู่ระหว่างดำเนินการ"
            return "❌ ยังไม่ได้ดำเนินการ" 
            
        df_combined["สถานะ"] = df_combined["สถานะดิบ"].apply(map_3_status)

    return df_combined

@st.cache_data(show_spinner=False)
def read_summary_sheet(file_bytes):
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    if "สรุป" not in wb.sheetnames: return pd.DataFrame()
    ws = wb["สรุป"]
    
    header_row = None
    for r in range(1, 20):
        row_vals = [str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else "" for c in range(1, ws.max_column + 1)]
        if "ประเภทงาน" in row_vals or "โครงการ" in row_vals:
            header_row = r
            break
            
    if not header_row: return pd.DataFrame()

    c_proj, c_note = None, None
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(header_row, c).value).strip()
        if "ประเภทงาน" in val or "โครงการ" in val: c_proj = c
        elif "หมายเหตุ" in val or "ปัญหา" in val: c_note = c

    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, c_proj).value if c_proj else None
        if not name or "รวม" in str(name): continue
        note = ws.cell(r, c_note).value if c_note else ""

        data.append({
            "โครงการ": str(name).strip(),
            "หมายเหตุ": str(note).strip() if note else "" 
        })
        
    df_summary = pd.DataFrame(data)
    if not df_summary.empty:
        # ตัดคำว่า (เพิ่มเติม) ออกเช่นเดียวกัน เพื่อให้จับคู่กับข้อมูลหลักได้
        df_summary["โครงการ"] = df_summary["โครงการ"].astype(str).str.replace(r'\s*\(เพิ่มเติม\)', '', regex=True)
        df_summary = df_summary.drop_duplicates(subset=['โครงการ'], keep='first')
        
    return df_summary

def render_html_table(df, cols):
    display_df = df[cols].copy()
    if "วงเงิน" in display_df.columns:
        display_df["วงเงิน"] = pd.to_numeric(display_df["วงเงิน"].astype(str).str.replace(',', ''), errors='coerce').fillna(0)
        display_df["วงเงิน"] = display_df["วงเงิน"].apply(lambda x: f"{x:,.0f}")
        
    return display_df.to_html(index=False, classes="styled-table", escape=False)

# ==================================================
# UI COMPONENTS
# ==================================================
def create_category_cards(summary_df):
    if summary_df.empty: return
    st.markdown("## 📂 สรุปตามหมวดงาน")
    cols = st.columns(4)

    for i, row in summary_df.reset_index(drop=True).iterrows():
        with cols[i % 4]:
            note = row.get("หมายเหตุ", "")
            total = row["จำนวนงาน"]
            done = row["งานแล้วเสร็จ"]
            pending = row["งานรอดำเนินการ"]
            percent = (done / total * 100) if total > 0 else 0
            
            st.markdown(f"""
            <div class="glass-card category-card">
                <div class="category-number">{str(i+1).zfill(2)}</div>
                <div class="category-title">{row["โครงการ"]}</div>
                <div class="category-value"><b>{total}</b> งานทั้งหมด</div>
                <div class="category-value">
                    <span style="color:#0BFF8D; font-weight:700;">✅ เสร็จ {done} งาน ({percent:.1f}%)</span>
                </div>
                <div style="width: 100%; background-color: #E2E8F0; border-radius: 999px; height: 6px; margin: 6px 0 12px 0;">
                    <div style="background-color: #0BFF8D; height: 6px; border-radius: 999px; width: {percent}%;"></div>
                </div>
                <div class="category-value"><span style="color:#D58A19; font-weight:700;">⏳ คงเหลือ {pending} งาน</span></div>
                <div class="category-note">{note}</div>
            </div>
            """, unsafe_allow_html=True)

def create_pea_chart(summary_df):
    if summary_df.empty: return
    def wrap_label(text):
        text = str(text)
        if len(text) > 20:
            if " " in text:
                return "<br>".join(textwrap.wrap(text, width=20))
            elif "/" in text:
                return text.replace("/", "/<br>")
            else:
                mid = len(text) // 2
                return text[:mid] + "<br>" + text[mid:]
        return text

    wrapped_x = summary_df["โครงการ"].apply(wrap_label)

    fig = go.Figure()
    
    fig.add_bar(
        x=wrapped_x, y=summary_df["จำนวนงาน"], name="จำนวนงานทั้งหมด", 
        marker_color="#3730A3", text=summary_df["จำนวนงาน"], textposition="outside", cliponaxis=False,
        marker_line_width=0
    )
    fig.add_bar(
        x=wrapped_x, y=summary_df["งานแล้วเสร็จ"], name="งานแล้วเสร็จ", 
        marker_color="#0BFF8D", text=summary_df["งานแล้วเสร็จ"], textposition="outside", cliponaxis=False,
        marker_line_width=0
    )

    fig.update_layout(
        barmode="group", 
        height=500, 
        paper_bgcolor="rgba(0,0,0,0)", 
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Sarabun, sans-serif", size=14, color="#1E293B"),
        legend=dict(orientation="h", y=1.12, font=dict(size=14, color="#1E293B")), 
        margin=dict(l=20, r=20, t=40, b=80), 
        xaxis=dict(showgrid=False, tickangle=0), 
        yaxis=dict(showgrid=True, gridcolor="#E2E8F0", gridwidth=1)
    )
    st.plotly_chart(fig, use_container_width=True)

# ==================================================
# PAGE 1: GOOGLE SHEETS CONNECT PAGE
# ==================================================
if st.session_state.app_state == 'connect':
    st.markdown("""
        <style>
        [data-testid="collapsedControl"], [data-testid="stSidebar"] { display: none !important; }
        
        @keyframes liquidGradient {
            0% { background-position: 0% 50%; }
            50% { background-position: 100% 50%; }
            100% { background-position: 0% 50%; }
        }
        
        [data-testid="stAppViewContainer"] {
            background: linear-gradient(-45deg, #FF9A9E, #FECFEF, #A1C4FD, #E0C3FC) !important;
            background-size: 400% 400% !important;
            animation: liquidGradient 12s ease infinite !important;
        }

        .liquid-header-box {
            background: rgba(255, 255, 255, 0.4);
            backdrop-filter: blur(15px);
            -webkit-backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.7);
            border-bottom: 1px solid rgba(255, 255, 255, 0.3);
            border-right: 1px solid rgba(255, 255, 255, 0.3);
            border-radius: 30px;
            padding: 40px;
            box-shadow: 0 8px 32px 0 rgba(31, 38, 135, 0.15);
            text-align: center;
            margin-bottom: 30px;
        }

        .upload-title {
            color: #4A4A4A;
            font-size: 42px;
            font-weight: 800;
            margin-bottom: 10px;
            text-shadow: 1px 1px 3px rgba(255,255,255,0.7);
        }
        
        .upload-subtitle {
            color: #4A4A4A;
            font-size: 18px;
            font-weight: 500;
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("<br><br>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div class="liquid-header-box">
            <div class="upload-title">📊 PEA Executive Dashboard</div>
            <div class="upload-subtitle">ระบบรายงานสรุปผลแบบเชื่อมต่อ Google Sheets (Private)</div>
        </div>
        """, unsafe_allow_html=True)
        
        url_input = st.text_input("🔗 วางลิงก์ Google Sheets (Private) ที่นี่:", placeholder="https://docs.google.com/spreadsheets/d/...")
        
        if st.button("🚀 โหลดข้อมูล", use_container_width=True):
            if not url_input:
                st.warning("⚠️ กรุณาวางลิงก์ Google Sheets ก่อนครับ")
            else:
                with st.spinner('กำลังใช้ API Key ดึงข้อมูลและประมวลผล (อาจใช้เวลาสักครู่)...'):
                    file_bytes = fetch_private_gsheet_excel(url_input)
                    
                    if file_bytes is not None:
                        summary_df_notes = read_summary_sheet(file_bytes)
                        file_bytes.seek(0) 
                        all_df = get_all_excel_data(file_bytes)
                        
                        if all_df.empty:
                            st.warning("❌ ไม่พบข้อมูลรายละเอียดโครงการในชีตย่อยเลย กรุณาตรวจสอบรูปแบบไฟล์")
                        else:
                            st.session_state.gsheet_url = url_input
                            st.session_state.summary_df = summary_df_notes 
                            st.session_state.all_df = all_df
                            st.session_state.app_state = 'dashboard'
                            st.rerun()

# ==================================================
# PAGE 2 & 3: DASHBOARD PAGE
# ==================================================
elif st.session_state.app_state == 'dashboard':
    all_df = st.session_state.all_df.copy()

    # --- ดึง 'ปี' โดยให้ความสำคัญกับ "ชื่อชีต" เป็นอันดับแรก ---
    if "ปี" not in all_df.columns:
        # 1. ลองดึงจากแหล่งที่มา (ชื่อชีต) ก่อน
        all_df["ปี"] = all_df["แหล่งที่มา (ชีต)"].apply(extract_year_from_text)
        
        # 2. ถ้าดึงจากชื่อชีตไม่ได้ (ได้ค่าว่าง) ค่อยไปหาในชื่อโครงการหลัก
        mask_empty = all_df["ปี"] == ""
        all_df.loc[mask_empty, "ปี"] = all_df.loc[mask_empty, "โครงการหลัก"].apply(extract_year_from_text)
        
        # 3. ถ้าหาไม่ได้เลยให้ใส่ "ไม่ระบุ"
        all_df["ปี"] = all_df["ปี"].replace("", "ไม่ระบุ")

    # ---------------------------------------------
    # SIDEBAR
    # ---------------------------------------------
    st.sidebar.markdown("### ⚙️ การจัดการข้อมูล")
    
    if st.sidebar.button("🔄 ดึงข้อมูลล่าสุด", type="primary", use_container_width=True):
        with st.spinner('กำลังใช้ API Key ดึงข้อมูลล่าสุด...'):
            file_bytes = fetch_private_gsheet_excel(st.session_state.gsheet_url)
            if file_bytes:
                st.session_state.summary_df = read_summary_sheet(file_bytes)
                file_bytes.seek(0)
                st.session_state.all_df = get_all_excel_data(file_bytes)
                st.cache_data.clear() 
                st.rerun()

    if st.sidebar.button("⬅️ เปลี่ยนลิงก์ Google Sheets", use_container_width=True):
        st.session_state.app_state = 'connect'
        st.session_state.summary_df = None
        st.session_state.all_df = None
        st.session_state.gsheet_url = ""
        st.rerun()
        
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 🔎 ตัวกรองข้อมูล")
    
    available_years = sorted(all_df["ปี"].unique().tolist())
    if "ไม่ระบุ" in available_years:
        available_years.remove("ไม่ระบุ")
        available_years.append("ไม่ระบุ")
    available_years.insert(0, "ทุกปี")
    
    selected_year = st.sidebar.selectbox("📅 เลือกปีงบประมาณ:", available_years)

    if selected_year == "ทุกปี":
        filtered_all_df = all_df.copy()
    else:
        filtered_all_df = all_df[all_df["ปี"] == selected_year].copy()

    # --- ส่วนที่เพิ่มใหม่: จัดกลุ่ม SPP ให้เป็นหมวดเดียวกัน ---
    def group_project_category(name):
        name_str = str(name).strip()
        if name_str.upper().startswith("SPP"):
            return "SPP" # ดึงทุกอย่างที่ขึ้นต้นด้วย SPP มารวมในชื่อนี้
        return name_str

    filtered_all_df["หมวดงาน_Dropdown"] = filtered_all_df["โครงการหลัก"].apply(group_project_category)
    # ----------------------------------------------------

    def generate_dynamic_summary(df):
        if df.empty: return pd.DataFrame()
        
        # เปลี่ยนการ Group By จาก 'โครงการหลัก' เป็น 'หมวดงาน_Dropdown'
        summary = df.groupby('หมวดงาน_Dropdown').agg({
            'ชื่องาน': 'count',
            'สถานะ': lambda x: (x == '✅ แล้วเสร็จ').sum()
        }).reset_index()
        
        summary.rename(columns={
            'หมวดงาน_Dropdown': 'โครงการ',
            'ชื่องาน': 'จำนวนงาน',
            'สถานะ': 'งานแล้วเสร็จ'
        }, inplace=True)
        
        summary['งานรอดำเนินการ'] = summary['จำนวนงาน'] - summary['งานแล้วเสร็จ']
        
        orig_sum = st.session_state.summary_df
        if orig_sum is not None and not orig_sum.empty and 'โครงการ' in orig_sum.columns:
            notes_dict = dict(zip(orig_sum['โครงการ'], orig_sum['หมายเหตุ']))
            
            # สร้างคำอธิบายสำหรับกลุ่ม SPP โดยดึงชื่อโครงการย่อยมาต่อกัน
            def get_note(proj):
                if proj == "SPP":
                    spp_list = df[df["หมวดงาน_Dropdown"] == "SPP"]["โครงการหลัก"].unique()
                    spp_list_clean = [s for s in spp_list if str(s).strip() != ""]
                    return f"รวมโครงการ: {', '.join(spp_list_clean)}"
                return notes_dict.get(proj, "")
                
            summary['หมายเหตุ'] = summary['โครงการ'].apply(get_note)
        else:
            summary['หมายเหตุ'] = ""
            
        return summary.sort_values(by='จำนวนงาน', ascending=False).reset_index(drop=True)

    filtered_summary_df = generate_dynamic_summary(filtered_all_df)

    st.sidebar.markdown("### 🎯 เจาะลึกรายละเอียดงาน")
    main_projects = filtered_summary_df["โครงการ"].dropna().astype(str).unique().tolist()
    main_projects = [p for p in main_projects if p.strip() and p.lower() != "nan"]
    main_projects.insert(0, "-- ภาพรวมทั้งหมด --")
    
    selected_main_proj = st.sidebar.selectbox("📌 เลือกหมวดงาน:", main_projects)

    # ---------------------------------------------
    # RENDER OVERVIEW (หน้าที่ 2)
    # ---------------------------------------------
    if selected_main_proj == "-- ภาพรวมทั้งหมด --":
        st.markdown("""
        <div class="pea-header">
            <div class="pea-title">Executive Dashboard</div>
            <div class="pea-subtitle">รายงานสรุปผลการดำเนินงานแบบเรียลไทม์ (Private Data)</div>
        </div>
        """, unsafe_allow_html=True)

        total_all = filtered_summary_df["จำนวนงาน"].sum() if not filtered_summary_df.empty else 0
        done_all = filtered_summary_df["งานแล้วเสร็จ"].sum() if not filtered_summary_df.empty else 0
        
        percent_all = (done_all / total_all * 100) if total_all > 0 else 0
        
        budget_series = filtered_all_df["วงเงิน"].astype(str).str.replace(',', '').str.replace(' ', '')
        budget_sum = pd.to_numeric(budget_series, errors='coerce').fillna(0).sum()

        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f'''
            <div class="glass-card">
                <div class="kpi-title">📌 ปริมาณงานทั้งหมด {f"({selected_year})" if selected_year != "ทุกปี" else ""}</div>
                <div class="kpi-value" style="color: #3730A3;">{total_all:,} <span style="font-size:16px; color:#94A3B8;">งาน</span></div>
            </div>''', unsafe_allow_html=True)
        with col2:
            st.markdown(f'''
            <div class="glass-card">
                <div class="kpi-title">✅ งานแล้วเสร็จ</div>
                <div class="kpi-value" style="color: #0BFF8D;">{done_all:,} <span style="font-size:16px; color:#94A3B8;">งาน</span> <span style="font-size:24px; color:#059669;">({percent_all:.1f}%)</span></div>
                <div style="width: 100%; background-color: #E2E8F0; border-radius: 999px; height: 8px; margin-top: 12px;">
                    <div style="background-color: #0BFF8D; height: 8px; border-radius: 999px; width: {percent_all}%;"></div>
                </div>
            </div>''', unsafe_allow_html=True)
        with col3:
            st.markdown(f'''
            <div class="glass-card">
                <div class="kpi-title">💰 วงเงินก่อสร้างรวม</div>
                <div class="kpi-value" style="color: #D97706;">{budget_sum:,.0f} <span style="font-size:16px; color:#94A3B8;">บาท</span></div>
            </div>''', unsafe_allow_html=True)

        if not filtered_summary_df.empty:
            st.markdown('<div class="glass-card">', unsafe_allow_html=True)
            create_pea_chart(filtered_summary_df)
            st.markdown('</div>', unsafe_allow_html=True)
            create_category_cards(filtered_summary_df)
        else:
            st.warning(f"⚠️ ไม่พบข้อมูลสำหรับปี: {selected_year}")

    # ---------------------------------------------
    # RENDER DRILL-DOWN (หน้าที่ 3)
    # ---------------------------------------------
    else:
        st.markdown(f"## 🔎 เจาะลึกรายละเอียดงาน: <span style='color:#3730A3;'>{selected_main_proj}</span>", unsafe_allow_html=True)
        
        # เปลี่ยนมาใช้การกรองจากคอลัมน์ หมวดงาน_Dropdown แทน
        detail_df = filtered_all_df[filtered_all_df["หมวดงาน_Dropdown"] == selected_main_proj].copy()

        if detail_df.empty:
            st.info(f"ไม่พบข้อมูลย่อยของ '{selected_main_proj}'")
        else:
            # เอาบรรทัด detail_df["โครงการหลัก"] = selected_main_proj ออก
            # เพื่อให้ตารางด้านล่างยังคงแสดงชื่อโปรเจกต์ย่อย (เช่น SPP (SCOD 70)) ได้อย่างชัดเจน
            pass 

            pie_data = detail_df["สถานะ"].value_counts().reset_index()
            pie_data.columns = ["Status", "Count"]

            color_map = {
                "✅ แล้วเสร็จ": "#0BFF8D",            
                "⏳ อยู่ระหว่างดำเนินการ": "#D58A19", 
                "❌ ยังไม่ได้ดำเนินการ": "#F43F5E"    
            }
            colors = [color_map.get(s, "#94A3B8") for s in pie_data["Status"]]

            sub_budget_series = detail_df["วงเงิน"].astype(str).str.replace(',', '').str.replace(' ', '')
            sub_budget = pd.to_numeric(sub_budget_series, errors='coerce').fillna(0).sum()

            st.markdown(f"<h4 style='color:#475569; margin-top:20px;'>📊 สัดส่วนสถานะงาน และข้อมูลวงเงินเฉพาะโครงการ</h4>", unsafe_allow_html=True)
            
            scol1, scol2, scol3 = st.columns([1, 1, 1.5])
            with scol1:
                total_sub_tasks = len(detail_df)
                st.markdown(f'''
                <div class="glass-card" style="border-left: 6px solid #3730A3;">
                    <p class="kpi-title" style="margin:0;">📌 งานย่อยทั้งหมด</p>
                    <p class="kpi-value" style="color:#0F172A; margin: 5px 0 0 0;">{total_sub_tasks}</p>
                </div>
                ''', unsafe_allow_html=True)
            with scol2:
                done_count = len(detail_df[detail_df["สถานะ"] == "✅ แล้วเสร็จ"])
                st.markdown(f'''
                <div class="glass-card" style="border-left: 6px solid #0BFF8D;">
                    <p class="kpi-title" style="margin:0;">✅ เสร็จสมบูรณ์แล้ว</p>
                    <p class="kpi-value" style="color:#0BFF8D; margin: 5px 0 0 0;">{done_count}</p>
                </div>
                ''', unsafe_allow_html=True)
            with scol3:
                st.markdown(f'''
                <div class="glass-card" style="border-left: 6px solid #D97706;">
                    <p class="kpi-title" style="margin:0;">💰 วงเงินเฉพาะหมวดนี้</p>
                    <p class="kpi-value" style="color:#D97706; margin: 5px 0 0 0;">{sub_budget:,.0f} ฿</p>
                </div>
                ''', unsafe_allow_html=True)

            col_chart, col_stat = st.columns([1, 1.5])
            with col_chart:
                fig_pie = go.Figure(data=[go.Pie(
                    labels=pie_data["Status"],
                    values=pie_data["Count"],
                    hole=0.55, 
                    marker=dict(colors=colors, line=dict(color='#FFFFFF', width=2)), 
                    textinfo='label+percent+value',
                    textposition='outside', 
                    hoverinfo='label+value'
                )])
                
                fig_pie.update_layout(
                    margin=dict(l=30, r=30, t=20, b=20), 
                    height=350,
                    showlegend=True,
                    legend=dict(orientation="h", y=-0.2, font=dict(color="#1E293B")),
                    font=dict(family="Sarabun, sans-serif", size=14, color="#1E293B"),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)"
                )
                st.plotly_chart(fig_pie, use_container_width=True)
                
            with col_stat:
                st.markdown("<br><br>", unsafe_allow_html=True)
                for _, row in pie_data.iterrows():
                    color_hex = color_map.get(row['Status'], "#1E293B")
                    st.markdown(f"**<span style='color:{color_hex}; font-size:18px;'>{row['Status']}</span> :** <span style='font-size:18px; font-weight:bold;'>{row['Count']}</span> งาน", unsafe_allow_html=True)

            st.markdown("---")
            st.markdown("### 📋 ตารางรายละเอียดงาน (เรียงตามกำหนดแล้วเสร็จ)")
            
            detail_df.rename(columns={"โครงการหลัก": "โครงการ (หมวดหลัก)"}, inplace=True)
            
            filter_col1, filter_col2 = st.columns(2)
            
            area_options = ["ทั้งหมด"] + sorted([str(x) for x in detail_df["พื้นที่"].unique() if pd.notna(x) and str(x).strip() not in ["", "-"]])
            owner_options = ["ทั้งหมด"] + sorted([str(x) for x in detail_df["ผู้รับผิดชอบ"].unique() if pd.notna(x) and str(x).strip() not in ["", "-"]])
            
            with filter_col1:
                selected_table_area = st.selectbox("📍 กรองข้อมูลตามพื้นที่ (เลือกได้ 1 พื้นที่):", options=area_options, index=0)
            with filter_col2:
                selected_table_owner = st.selectbox("👤 กรองข้อมูลตามผู้รับผิดชอบ (เลือกได้ 1 ผู้รับผิดชอบ):", options=owner_options, index=0)
                
            filtered_table_df = detail_df.copy()
            if selected_table_area != "ทั้งหมด":
                filtered_table_df = filtered_table_df[filtered_table_df["พื้นที่"].astype(str) == selected_table_area]
            if selected_table_owner != "ทั้งหมด":
                filtered_table_df = filtered_table_df[filtered_table_df["ผู้รับผิดชอบ"].astype(str) == selected_table_owner]
            
            display_cols = ["โครงการ (หมวดหลัก)", "ชื่องาน", "แหล่งที่มา (ชีต)", "พื้นที่", "จังหวัด", "ผู้รับผิดชอบ", "วงเงิน", "หมายเหตุ/ปัญหา", "Deadline"]

            df_done = filtered_table_df[filtered_table_df["สถานะ"] == "✅ แล้วเสร็จ"].sort_values(by="Sort_Index")
            st.markdown('<div class="table-header-done">✅ งานที่แล้วเสร็จ</div>', unsafe_allow_html=True)
            if not df_done.empty:
                st.markdown(render_html_table(df_done, display_cols), unsafe_allow_html=True)
            else:
                st.info("ไม่มีข้อมูลที่ตรงกับเงื่อนไขในสถานะนี้")

            df_prog = filtered_table_df[filtered_table_df["สถานะ"] == "⏳ อยู่ระหว่างดำเนินการ"].sort_values(by="Sort_Index")
            st.markdown('<div class="table-header-prog">⏳ งานที่อยู่ระหว่างดำเนินการ</div>', unsafe_allow_html=True)
            if not df_prog.empty:
                st.markdown(render_html_table(df_prog, display_cols), unsafe_allow_html=True)
            else:
                st.info("ไม่มีข้อมูลที่ตรงกับเงื่อนไขในสถานะนี้")

            df_not = filtered_table_df[filtered_table_df["สถานะ"] == "❌ ยังไม่ได้ดำเนินการ"].sort_values(by="Sort_Index")
            st.markdown('<div class="table-header-not">❌ งานที่ยังไม่ได้ดำเนินการ</div>', unsafe_allow_html=True)
            if not df_not.empty:
                st.markdown(render_html_table(df_not, display_cols), unsafe_allow_html=True)
            else:
                st.info("ไม่มีข้อมูลที่ตรงกับเงื่อนไขในสถานะนี้")
